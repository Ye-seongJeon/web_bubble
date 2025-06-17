import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Weight constants (adjust these to change the weight of each component)
CONSTRAINTS_WEIGHT = 0.15
VOLUME_WEIGHT = 0.075
BB_WEIGHT = 0.075
SURFACE_VIEW_WEIGHT = 0.20
SECTION_VIEW_WEIGHT = 0.20
CHAMFER_DISTANCE_WEIGHT = 0.15
AAD_WEIGHT = 0.15

def load_json(file_path):
    with open(file_path, 'r') as f:
        return json.load(f)

def calculate_score(student_data, stp_results, stl_results, constraints_results):
    constraints_score = next(iter(constraints_results.values()))['score']  # Extract the 'score' from the inner dict
    volume_score = student_data.get('volume_score', 0)
    bb_score = student_data.get('bb_score', 0)
    surface_view_score = student_data.get('surface_view_score', 0)
    section_view_score = student_data.get('section_view_score', 0)
    
    chamfer_distance = stl_results.get('ChamferDistance', 0)
    max_chamfer = max((result.get('ChamferDistance', 0) for result in stl_results.values() if isinstance(result, dict)), default=1)
    chamfer_score = 100 * (1 - chamfer_distance / max_chamfer) if max_chamfer != 0 else 0
    
    aad_similarity = stl_results.get('AADSimilarity')
    aad_score = aad_similarity if aad_similarity is not None else 0

    total_score = (
        CONSTRAINTS_WEIGHT * constraints_score +
        VOLUME_WEIGHT * volume_score +
        BB_WEIGHT * bb_score +
        SURFACE_VIEW_WEIGHT * surface_view_score +
        SECTION_VIEW_WEIGHT * section_view_score +
        CHAMFER_DISTANCE_WEIGHT * chamfer_score +
        AAD_WEIGHT * aad_score
    )

    return total_score

def create_excel_report(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results of Students"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Add title
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = "3D CAD Automatic Grading System Results"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    headers = [
        "Name", "Total Score", "Check for File Validity", "Constraints Check",
        "volume_score", "bb_score", "surface_view_score", "section_view_score",
        "ChamferDistance", "EarthMoversDistance", "PointMeshDistance", "HausdorffDistance", "AADSimilarity"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border

    # Write data
    for row, (name, data) in enumerate(results.items(), start=3):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(data['Total Score'], 2))
        ws.cell(row=row, column=3, value="TRUE")
        ws.cell(row=row, column=4, value=data['Constraints Check'])
        ws.cell(row=row, column=5, value=round(data['volume_score'], 2))
        ws.cell(row=row, column=6, value=round(data['bb_score'], 2))
        ws.cell(row=row, column=7, value=round(data['surface_view_score'], 2))
        ws.cell(row=row, column=8, value=round(data['section_view_score'], 2))
        ws.cell(row=row, column=9, value=round(data['ChamferDistance'], 4))
        ws.cell(row=row, column=10, value=round(data['EarthMoversDistance'], 4))
        ws.cell(row=row, column=11, value=round(data['PointMeshDistance'], 4))
        ws.cell(row=row, column=12, value=round(data['HausdorffDistance'], 4))
        ws.cell(row=row, column=13, value=round(data['AADSimilarity'], 2) if data['AADSimilarity'] is not None else None)

        # Apply styles
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.alignment = centered_alignment
            cell.border = border

    # Add weight information
    weight_info = [
        "Weight Information:",
        f"Constraints: {CONSTRAINTS_WEIGHT}",
        f"Volume: {VOLUME_WEIGHT}",
        f"Bounding Box: {BB_WEIGHT}",
        f"Surface View: {SURFACE_VIEW_WEIGHT}",
        f"Section View: {SECTION_VIEW_WEIGHT}",
        f"Chamfer Distance: {CHAMFER_DISTANCE_WEIGHT}",
        f"AAD: {AAD_WEIGHT}"
    ]
    for i, info in enumerate(weight_info, start=len(results) + 4):
        ws.cell(row=i, column=1, value=info)

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(r"code_ver4.0.0/paper_testcase/results/student_results.xlsx")

def main():
    stp_results = load_json(r"code_ver4.0.0/paper_testcase/results/hw1_stp_results.json")
    stl_results = load_json(r"code_ver4.0.0/paper_testcase/results/hw1_stl_results.json")

    results = {}
    for i in range(1, 16):
        student_id = f"hw1_s{i}"
        constraints_file = f"code_ver4.0.0/paper_testcase/hw1/hw1_constraints_s{i}.json"
        
        if os.path.exists(constraints_file):
            constraints_results = load_json(constraints_file)
        else:
            print(f"Warning: Constraints file not found for {student_id}.")
            continue  # Skip this student if constraints file is not found

        if student_id in stp_results and student_id in stl_results:
            total_score = calculate_score(stp_results[student_id], stp_results, stl_results[student_id], constraints_results)
            
            results[f"Student{i:02d}"] = {
                "Total Score": total_score,
                "Constraints Check": next(iter(constraints_results.values()))['score'],  # Extract the 'score' from the inner dict
                "volume_score": stp_results[student_id].get('volume_score', 0),
                "bb_score": stp_results[student_id].get('bb_score', 0),
                "surface_view_score": stp_results[student_id].get('surface_view_score', 0),
                "section_view_score": stp_results[student_id].get('section_view_score', 0),
                "ChamferDistance": stl_results[student_id].get('ChamferDistance', 0),
                "EarthMoversDistance": stl_results[student_id].get('EarthMoversDistance', 0),
                "PointMeshDistance": stl_results[student_id].get('PointMeshDistance', 0),
                "HausdorffDistance": stl_results[student_id].get('HausdorffDistance', 0),
                "AADSimilarity": stl_results[student_id].get('AADSimilarity')
            }
        else:
            print(f"Warning: Data not found for {student_id}")

    create_excel_report(results)
    print("Excel report generated: student_results.xlsx")

if __name__ == "__main__":
    main()